import os
import csv
import json
from datetime import datetime
from urllib.request import urlretrieve
from openpyxl import load_workbook

def load_json_data(file_path):
    with open(file_path) as file:
        data = json.load(file)
    return data

def load_xlsx_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = {}
    for row in sheet.iter_rows(values_only=True):
        try:
            code, country = row
            data[int(code)] = country
        except (ValueError, TypeError):
            continue
    return data

rating_mapping = {
    "Average": "Average",
    "Bardzo dobrze": "Very Good",
    "Bueno": "Good",
    "Eccellente": "Excellent",
    "Excelente": "Excellent",
    "Excellent": "Excellent",
    "Good": "Good",
    "Muito Bom": "Very Good",
    "Muy Bueno": "Very Good",
    "Not rated": "Not rated",
    "Poor": "Poor",
    "Skvělá volba": "Excellent",
    "Skvělé": "Excellent",
    "Terbaik": "Excellent",
    "Velmi dobré": "Very Good",
    "Very Good": "Very Good"
}

# Extract the following fields and store the data as restaurants.csv.
def extract_restaurants(data, country_codes):
    restaurants = []
    for item in data:
        for restaurant_data in item['restaurants']:
            restaurant = restaurant_data['restaurant']
            country_code = restaurant['location']['country_id']
            country = country_codes.get(country_code, 'Unknown')
            rating_text = restaurant['user_rating']['rating_text']
            rating_category = rating_mapping.get(rating_text, 'Unknown')
            restaurants.append({
                'Restaurant Id': restaurant['R']['res_id'] if restaurant['R']['res_id'] else 'NA',
                'Restaurant Name': restaurant['name'] if restaurant['name'] else 'NA',
                'Country': country if country else 'NA',
                'City': restaurant['location']['city'] if restaurant['location']['city'] else 'NA',
                'User Rating Votes': restaurant['user_rating']['votes'] if restaurant['user_rating']['votes'] else 'NA',
                'User Aggregate Rating': float(restaurant['user_rating']['aggregate_rating']) if restaurant['user_rating']['aggregate_rating'] else 'NA',
                'Rating Category': rating_category,
                'Cuisines': restaurant['cuisines'] if restaurant['cuisines'] else 'NA'
            })
    return restaurants

def write_csv(file_path, data, fieldnames):
    with open(file_path, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    print(f"Data written to {file_path}")

# Extract the list of restaurants that have past event in the month of April 2019 and store the data as restaurant_events.csv.
def extract_restaurant_events(data):
    events = []
    for item in data:
        for restaurant_data in item['restaurants']:
            restaurant = restaurant_data['restaurant']
            if 'zomato_events' in restaurant:
                for event_data in restaurant['zomato_events']:
                    event = event_data['event']
                    start_date = datetime.strptime(event['start_date'], '%Y-%m-%d')
                    if start_date.month == 4 and start_date.year == 2019:
                        events.append({
                            'Event Id': event['event_id'] if event['event_id'] else 'NA',
                            'Restaurant Id': restaurant['R']['res_id'] if restaurant['R']['res_id'] else 'NA',
                            'Restaurant Name': restaurant['name'] if restaurant['name'] else 'NA',
                            'Photo URL': event['photos'][0]['photo']['url'] if event['photos'] else 'NA',
                            'Event Title': event['title'] if event['title'] else 'NA',
                            'Event Start Date': event['start_date'] if event['start_date'] else 'NA',
                            'Event End Date': event['end_date'] if event['end_date'] else 'NA'
                        })
    return events

# From the dataset (restaurant_data.json), determine the threshold for the different rating text based on aggregate rating.
def get_rating_thresholds(data):
    ratings = []
    for item in data:
        for restaurant_data in item['restaurants']:
            restaurant = restaurant_data['restaurant']
            if 'user_rating' in restaurant:
                ratings.append(float(restaurant['user_rating']['aggregate_rating']))
    
    ratings = sorted(set(ratings))
    
    thresholds = {}
    if len(ratings) >= 5:
        thresholds = {
            'Excellent': (ratings[-1], ratings[-1]),
            'Very Good': (ratings[-2], ratings[-1]),
            'Good': (ratings[-3], ratings[-2]),
            'Average': (ratings[-4], ratings[-3]),
            'Poor': (ratings[0], ratings[-4])
        }
    elif len(ratings) == 4:
        thresholds = {
            'Excellent': (ratings[-1], ratings[-1]),
            'Very Good': (ratings[-2], ratings[-1]),
            'Good': (ratings[-3], ratings[-2]),
            'Average': (ratings[0], ratings[-3])
        }
    elif len(ratings) == 3:
        thresholds = {
            'Excellent': (ratings[-1], ratings[-1]),
            'Very Good': (ratings[-2], ratings[-1]),
            'Good': (ratings[0], ratings[-2])
        }
    elif len(ratings) == 2:
        thresholds = {
            'Excellent': (ratings[-1], ratings[-1]),
            'Very Good': (ratings[0], ratings[-1])
        }
    elif len(ratings) == 1:
        thresholds = {
            'Excellent': (ratings[0], ratings[0])
        }
    
    return thresholds

def ensure_folder_exists(folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

if __name__ == '__main__':
    data_folder = 'data'
    ensure_folder_exists(data_folder)
    
    url = 'https://raw.githubusercontent.com/Papagoat/brain-assessment/main/restaurant_data.json'
    urlretrieve(url, os.path.join(data_folder, 'restaurant_data.json'))
    data = load_json_data(os.path.join(data_folder, 'restaurant_data.json'))
    
    country_codes_url = 'https://github.com/Papagoat/brain-assessment/blob/main/Country-Code.xlsx?raw=true'
    urlretrieve(country_codes_url, os.path.join(data_folder, 'Country-Code.xlsx'))
    country_codes = load_xlsx_data(os.path.join(data_folder, 'Country-Code.xlsx'))
    
    restaurants = extract_restaurants(data, country_codes)
    write_csv(os.path.join(data_folder, 'restaurants.csv'), restaurants, fieldnames=restaurants[0].keys())
    
    events = extract_restaurant_events(data)
    write_csv(os.path.join(data_folder, 'restaurant_events.csv'), events, fieldnames=events[0].keys())
    
    thresholds = get_rating_thresholds(data)
    print("User Aggregate Rating Thresholds:")
    for rating, (min_val, max_val) in thresholds.items():
        print(f"{rating}: {min_val} - {max_val}")
