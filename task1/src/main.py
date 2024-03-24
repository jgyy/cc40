import os
import csv
import json
from datetime import datetime
from urllib.request import urlretrieve
from openpyxl import load_workbook

# not all ratings are in English
rating_mapping = {
    "Average": "Average",
    "Bardzo dobrze": "Very Good",
    "Bueno": "Good",
    "Eccellente": "Excellent",
    "Excelente": "Excellent",
    "Excellent": "Excellent",
    "Good": "Good",
    "Muito Bom": "Very Good",
    "Muy Bueno": "Very Good",
    "Not rated": "Not rated",
    "Poor": "Poor",
    "Skvělá volba": "Excellent",
    "Skvělé": "Excellent",
    "Terbaik": "Excellent",
    "Velmi dobré": "Very Good",
    "Very Good": "Very Good"
}


# Loads data from a JSON file
def load_json_data(file_path):
    with open(file_path) as file:
        data = json.load(file)
    return data


# xlsx is a binary fine hence need openpyxl library
def load_xlsx_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = {}
    for row in sheet.iter_rows(values_only=True):
        try:
            code, country = row
            data[int(code)] = country
        except (ValueError, TypeError):
            continue
    return data


# Extract the following fields and store the data as restaurants.csv.
def extract_restaurants(data, country_codes):
    restaurants = []

    for item in data:
        for restaurant_data in item['restaurants']:
            restaurant = restaurant_data['restaurant']
            country_code = restaurant['location']['country_id']
            country = country_codes.get(country_code, 'Unknown')
            rating_text = restaurant['user_rating']['rating_text']
            rating_category = rating_mapping.get(rating_text, 'Unknown')
            restaurant_id = restaurant['R'].get('res_id', 'NA')
            restaurant_name = restaurant.get('name', 'NA')
            city = restaurant['location'].get('city', 'NA')
            user_rating_votes = restaurant['user_rating'].get('votes', 'NA')
            user_rating = 'NA'
            if restaurant['user_rating']['aggregate_rating']:
                user_rating = restaurant['user_rating']['aggregate_rating']
                user_rating = float(user_rating)
            cuisines = restaurant.get('cuisines', 'NA')
            restaurants.append({
                'Restaurant Id': restaurant_id,
                'Restaurant Name': restaurant_name,
                'Country': country,
                'City': city,
                'User Rating Votes': user_rating_votes,
                'User Aggregate Rating': user_rating,
                'Rating Category': rating_category,
                'Cuisines': cuisines
            })

    return restaurants


# Writes the extracted data to a CSV file
def write_csv(file_path, data, fieldnames):
    with open(file_path, 'w', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    print(f"Data written to {file_path}")


# Extract the list of restaurants that have past event in the month
# of April 2019 and store the data as restaurant_events.csv.
def extract_restaurant_events(data):
    events = []
    for item in data:
        for restaurant_data in item['restaurants']:
            restaurant = restaurant_data['restaurant']
            if 'zomato_events' in restaurant:
                for event_data in restaurant['zomato_events']:
                    event = event_data['event']
                    start_date = datetime.strptime(
                        event.get('start_date', ''), '%Y-%m-%d')
                    if start_date.month == 4 and start_date.year == 2019:
                        event_id = event.get('event_id', 'NA')
                        res_id = restaurant.get('R', {}).get('res_id', 'NA')
                        restaurant_name = restaurant.get('name', 'NA')
                        if event['photos']:
                            photo_url = event['photos'][0]['photo']['url']
                        else:
                            photo_url = 'NA'
                        event_title = event.get('title', 'NA')
                        event_start_date = event.get('start_date', 'NA')
                        event_end_date = event.get('end_date', 'NA')
                        events.append({
                            'Event Id': event_id,
                            'Restaurant Id': res_id,
                            'Restaurant Name': restaurant_name,
                            'Photo URL': photo_url,
                            'Event Title': event_title,
                            'Event Start Date': event_start_date,
                            'Event End Date': event_end_date
                        })
    return events


# From the dataset (restaurant_data.json), determine the threshold
# for the different rating text based on aggregate rating.
def get_rating_thresholds(data):
    ratings = sorted(set(float(restaurant['user_rating']['aggregate_rating'])
                         for item in data
                         for restaurant_data in item['restaurants']
                         for restaurant in [restaurant_data['restaurant']]
                         if 'user_rating' in restaurant))

    ratings = sorted(set(ratings))
    thresholds = {}
    if len(ratings) >= 5:
        thresholds = {
            'Excellent': (ratings[-1], ratings[-1]),
            'Very Good': (ratings[-2], ratings[-1]),
            'Good': (ratings[-3], ratings[-2]),
            'Average': (ratings[-4], ratings[-3]),
            'Poor': (ratings[0], ratings[-4])
        }
    elif len(ratings) == 4:
        thresholds = {
            'Excellent': (ratings[-1], ratings[-1]),
            'Very Good': (ratings[-2], ratings[-1]),
            'Good': (ratings[-3], ratings[-2]),
            'Average': (ratings[0], ratings[-3])
        }
    elif len(ratings) == 3:
        thresholds = {
            'Excellent': (ratings[-1], ratings[-1]),
            'Very Good': (ratings[-2], ratings[-1]),
            'Good': (ratings[0], ratings[-2])
        }
    elif len(ratings) == 2:
        thresholds = {
            'Excellent': (ratings[-1], ratings[-1]),
            'Very Good': (ratings[0], ratings[-1])
        }
    elif len(ratings) == 1:
        thresholds = {
            'Excellent': (ratings[0], ratings[0])
        }
    return thresholds


def ensure_folder_exists(folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)


if __name__ == '__main__':
    data_folder = 'data'
    ensure_folder_exists(data_folder)

    url = ('https://raw.githubusercontent.com/Papagoat/brain-assessment/main/'
           'restaurant_data.json')
    urlretrieve(url, os.path.join(data_folder, 'restaurant_data.json'))
    data = load_json_data(os.path.join(data_folder, 'restaurant_data.json'))

    country_codes_url = ('https://github.com/Papagoat/brain-assessment/blob/'
                         'main/Country-Code.xlsx?raw=true')
    urlretrieve(country_codes_url,
                os.path.join(data_folder, 'Country-Code.xlsx'))
    country_codes = load_xlsx_data(
        os.path.join(data_folder, 'Country-Code.xlsx'))
    restaurants = extract_restaurants(data, country_codes)

    write_csv(os.path.join(data_folder, 'restaurants.csv'),
              restaurants, fieldnames=restaurants[0].keys())
    events = extract_restaurant_events(data)
    write_csv(os.path.join(data_folder, 'restaurant_events.csv'),
              events, fieldnames=events[0].keys())
    thresholds = get_rating_thresholds(data)
    print("User Aggregate Rating Thresholds:")
    for rating, (min_val, max_val) in thresholds.items():
        print(f"{rating}: {min_val} - {max_val}")
